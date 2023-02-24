# import numpy as np
# data=([[ 0.86466285, 0.76943895, 0.22678279], 
# [ 0.12452825, 0.54751384, 0.06499123],
# [ 0.06216566, 0.85045125, 0.92093862],
# [ 0.58401239, 0.93455057, 0.28972379]])

# # np.save('new_saved',data)  
# m=np.load('new_saved.npy')
# print(m)
import openpyxl

## opening the previously created xlsx file using 'load_workbook()' method
xlsx = openpyxl.load_workbook('sample.xlsx')

## getting the sheet to active
sheet = xlsx.active

## getting the reference of the cells which we want to get the data from
name = sheet['A1']
tag = sheet['B1']

#or in other way
name1 = sheet.cell(row=1,column=1)

## printing the values of cells
print(name.value)
print(tag.value)
print(name1.value)


